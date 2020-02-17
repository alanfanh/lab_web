# -*- coding: utf-8 -*-
import re
import os
import uuid
import xlrd, xlwt
from xlrd import xldate_as_tuple
from openpyxl import Workbook, load_workbook
from datetime import datetime
from werkzeug import secure_filename

try:
    from urlparse import urlparse, urljoin
except ImportError:
    from urllib.parse import urlparse, urljoin

# import PIL
# from PIL import Image
from flask import current_app, request, url_for, redirect, flash
from itsdangerous import BadSignature, SignatureExpired
from itsdangerous import TimedJSONWebSignatureSerializer as Serializer

from config import Operations
from . import db

def generate_token(user, operation, expire_in=None, **kwargs):
    s = Serializer(current_app.config['SECRET_KEY'], expire_in)
    data = {'id': user.id, 'operation': operation}
    data.update(**kwargs)
    return s.dumps(data)


def validate_token(user, token, operation, new_password=None):
    s = Serializer(current_app.config['SECRET_KEY'])

    try:
        data = s.loads(token)
    except (SignatureExpired, BadSignature):
        return False

    if operation != data.get('operation') or user.id != data.get('id'):
        return False

    if operation == Operations.CONFIRM:
        user.confirmed = True
    elif operation == Operations.RESET_PASSWORD:
        user.set_password(new_password)
    elif operation == Operations.CHANGE_EMAIL:
        new_email = data.get('new_email')
        if new_email is None:
            return False
        if User.query.filter_by(email=new_email).first() is not None:
            return False
        user.email = new_email
    else:
        return False

    db.session.commit()
    return True


def rename_file(old_filename):
    ext = os.path.splitext(old_filename)[1]
    new_filename = uuid.uuid4().hex + ext
    return new_filename


# def resize_image(image, filename, base_width):
    # filename, ext = os.path.splitext(filename)
    # img = Image.open(image)
    # if img.size[0] <= base_width:
        # return filename + ext
    # w_percent = (base_width / float(img.size[0]))
    # h_size = int((float(img.size[1]) * float(w_percent)))
    # img = img.resize((base_width, h_size), PIL.Image.ANTIALIAS)

    # filename += current_app.config['ALBUMY_PHOTO_SUFFIX'][base_width] + ext
    # img.save(os.path.join(current_app.config['ALBUMY_UPLOAD_PATH'], filename), optimize=True, quality=85)
    # return filename


def is_safe_url(target):
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ('http', 'https') and \
           ref_url.netloc == test_url.netloc


def redirect_back(default='main.index', **kwargs):
    print("*****request.args.get('next')=", request.args.get('next'))
    print("*****request.referrer=", request.referrer)
    print("*****kwargs=", **kwargs)
    print("*****default=", default)
    for target in request.args.get('next'), request.referrer:
        if not target:
            continue
        if is_safe_url(target):
            return redirect(target)


    return redirect(url_for(default, **kwargs))


def flash_errors(form):
    for field, errors in form.errors.items():
        for error in errors:
            flash(u"Error in the %s field - %s" % (
                getattr(form, field).label.text,
                error
            ))


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in current_app.config['ALLOWED_EXTENSIONS']

#读取excel    
def read_excel(path):
    filename=os.path.basename(path)
    if filename.endswith('xls'):
        worksheet = xlrd.open_workbook(path)
        sheet = worksheet.sheet_by_index(0)
        rows = sheet.nrows
        columns = sheet.ncols
        data = []
        head = sheet.row_values(0)
        data.append(head)
        for i in range(1, rows):
            row = sheet.row_values(i)
            if row:
                app = {}
                for j in range(len(row)):
                    ctype = sheet.cell(i, j).ctype
                    # 判断单元格数据类型,如果是date,则进行处理。
                    if ctype == 3:
                        date = datetime(*xldate_as_tuple(row[j], 0))
                        row[j] = date.strftime('%Y-%m-%d')
                    elif ctype == 2:
                        row[j] = int(row[j])
                    app[head[j]] = row[j]
                data.append(app)
    elif filename.endswith('xlsx'):
        worksheet = load_workbook(path)
        sheets = worksheet.get_sheet_names()
        sheet = worksheet.get_sheet_by_name(sheets[0])
        rows = sheet.rows
        columns = sheet.columns
        data = []
        head = [col.value for col in rows[0]]
        data.append(head)
        for row in rows:
            app = {}
            if row:
                for i in range(len(row)):
                    app[head[i]] = row[i].value
                data.append(app)
        # data.remove(data[0])
    return data



def write_excel(data,filename):
    """写excel"""
    print("**start write**")
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet('data',cell_overwrite_ok=True)
    # list_table_head = [data[0][0],data[0][1],data[0][2],data[0][3],data[0][4],data[0][5],data[0][6],data[0][7],data[0][8],data[0][9],data[0][10],data[0][11],data[0][12]]
    for row in range(len(data)):
        for col in range(len(data[row])):
            sheet.write(row,col,data[row][col])
    workbook.save(filename)    


#检查表格格式是否正确
def checkHead(fact,expe):
    if fact == expe:
        return True
    else:
        return False



#检查是否为正整数
def checkInt(data):
    if re.match(r'^\d+(\.0)?$',str(data)):
        return True
    else:
        return False


#检查是否为正整数和小数
def checkNumber(data):
    if re.match(r'^\d+\.?\d+$',str(data)):
        print(111)
        return True
    else:
        if checkInt(data):
            print(222)
            return True
        return False
    
    
#检查是否为数字字母和下划线
def checkType(data):
    if re.match(r'^[0-9A-Za-z_]+$', str(data)):
        print(str(data))
        return True
    else:
        return False

#检查是否为空
def checkEmpty(data):
    if str(data) == "":
        return True
    else:
        return False

#检查是否为数字字母中文和下划线
def checkType2(data):
    if re.match(r'^[0-9a-zA-Z_\u4e00-\u9fa5]+$', str(data)):
        return True
    else:
        return False

#检查是否是符合格式的日期
def checkDate(data):
    if re.match(r'^(\d{4})-(\d+)-(\d+)$', str(data)):
        try:
            if datetime.strptime(str(data), '%Y-%m-%d'):
                return True
        except Exception as e:
            return False
    elif re.match(r'^(\d{4})/(\d+)/(\d+)$', str(data)):
        try:
            if datetime.strptime(data,'%Y/%m/%d'):
                return True
        except Exception as e:
            return False
    return False

#检查日期2是否大于等于日期1
def checkCompareDate(data1,data2):
    if re.match(r'^(\d{4})-(\d+)-(\d+)$',str(data1)):
        data1 = datetime.strptime(data1,'%Y-%m-%d')
    elif re.match(r'^(\d{4})/(\d+)/(\d+)$',str(data1)):
        data1 = datetime.strptime(data1,'%Y/%m/%d')
    if re.match(r'^(\d{4})-(\d+)-(\d+)$',str(data2)):
        data2 = datetime.strptime(data2,'%Y-%m-%d')
    elif re.match(r'^(\d{4})/(\d+)/(\d+)$',str(data2)):
        data2 = datetime.strptime(data2,'%Y/%m/%d')
    print ('data2',data2)
    print ('data1',data1)
    if data2 < data1:
        return True
    else: 
        return False

#检查邮箱格式 
def checkEmail(data):
    if re.match(r'^[a-z0-9]+([._\\-]*[a-z0-9])*@([a-z0-9]+[-a-z0-9]*[a-z0-9]+.){1,63}[a-z0-9]+$',str(data)):
        return True
    else:
        return False

#检查长度
def checkLength(data,len1,len2):
    factlen = len(str(data))
    if  len1<=factlen<=len2:
        return True
    else:
        return False
